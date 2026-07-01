from flask import Flask, request, jsonify, g
from flask_cors import CORS
# from db import execute_query  # Oracle/psycopg2 execute_query commented out - using Supabase SDK only
from supabase_client import (
    select as sb_select,
    insert as sb_insert,
    update as sb_update,
    delete as sb_delete,
    select_distinct as sb_select_distinct,
    delete_all as sb_delete_all,
)
from config import config
import datetime
import sys
import logging
import os

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s: %(message)s',
    handlers=[
        logging.FileHandler('backend_debug.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev_secret_key')
CORS(app, resources={r"/api/*": {"origins": "*"}}, allow_headers=["Content-Type", "Authorization", "X-DB-User", "X-DB-Password", "X-DB-Type"])


@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"error": str(e)}), 500


@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    user = data.get('username')
    password = data.get('password')
    # db_type is always 'postgres' now - Oracle login removed
    # db_type = data.get('db_type', 'oracle')

    if not user or not password:
        return jsonify({"error": "Usuario y contraseña requeridos"}), 400

    if config.SUPABASE_URL is None or config.SUPABASE_SERVICE_ROLE_KEY is None:
        return jsonify({"error": "Supabase no está configurado para PostgreSQL"}), 500

    # Validar que el usuario coincide con PG_SCHEMA
    expected_user = (config.PG_SCHEMA or '').strip().lower()
    entered_user = (user or '').strip().lower()
    if not expected_user or entered_user != expected_user:
        logging.warning(f"Login fallido: usuario '{user}' no coincide con PG_SCHEMA '{config.PG_SCHEMA}'")
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    # Validar contraseña contra PG_PASSWORD
    if config.PG_PASSWORD:
        entered = (password or '').strip()
        expected = config.PG_PASSWORD.strip()
        logging.info(f"Login: PG_PASSWORD configurado, validando contraseña para user={user}")
        if entered != expected:
            logging.warning(f"Login fallido para {user}: contraseña no coincide con PG_PASSWORD")
            return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    logging.info(f"Login successful for user={user} using Supabase API")
    g.db_user = user
    g.db_password = password
    g.db_type = 'postgres'
    return jsonify({"status": "success", "user": user, "db_type": "postgres"})

    # ============================================================
    # ORACLE LOGIN - COMMENTED OUT (not in use)
    # ============================================================
    # try:
    #     g.db_user = user
    #     g.db_password = password
    #     g.db_type = db_type
    #     logging.info(f"Attempting login for user={user} on db_type={db_type}")
    #     test_query = "SELECT 1 FROM DUAL"
    #     execute_query(test_query)
    #     logging.info(f"Login successful for {user}")
    #     return jsonify({"status": "success", "user": user, "db_type": db_type})
    # except Exception as e:
    #     logging.error(f"Login failed for {user} on {db_type}: {str(e)}")
    #     return jsonify({"error": f"Error de conexión a la base de datos: {str(e)}"}), 401

@app.before_request
def before_request_func():
    from flask import g
    g.db_user = request.headers.get('X-DB-User')
    g.db_password = request.headers.get('X-DB-Password')
    # Always PostgreSQL/Supabase - Oracle removed
    g.db_type = 'postgres'
    logging.debug(f"[DB] db_type: postgres (Supabase)")


# is_postgres() kept for compatibility but always returns True
def is_postgres():
    return True


@app.route('/api/debug/pg', methods=['GET'])
def debug_pg():
    """Endpoint de diagnóstico: muestra la configuración de PostgreSQL sin revelar secretos."""
    return jsonify({
        "PG_HOST": config.PG_HOST,
        "PG_PORT": config.PG_PORT,
        "PG_DB": config.PG_DB,
        "PG_SCHEMA": config.PG_SCHEMA,
        "PG_PROJECT_REF": config.PG_PROJECT_REF,
        "PG_PASSWORD_configurado": bool(config.PG_PASSWORD),
        "PG_PASSWORD_longitud": len(config.PG_PASSWORD) if config.PG_PASSWORD else 0,
        "USANDO_SUPABASE": True,
        "SUPABASE_URL_configurada": bool(config.SUPABASE_URL),
        "SUPABASE_KEY_configurada": bool(config.SUPABASE_SERVICE_ROLE_KEY),
        "SUPABASE_SERVICE_ROLE_KEY_configurada": bool(config.SUPABASE_SERVICE_ROLE_KEY),
    })

@app.route('/api/debug/vacaciones', methods=['GET'])
def debug_vacaciones():
    from supabase_client import select as sb_select
    try:
        vacations = sb_select('VACACIONES')
        return jsonify({
            "count": len(vacations),
            "sample": vacations[:2] if vacations else [],
            "error": None
        })
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500

@app.route('/api/debug/connection', methods=['GET'])
def debug_connection():
    """Diagnóstico completo de la conexión a Supabase."""
    import traceback
    from supabase_client import SUPABASE_URL, SUPABASE_KEY, SCHEMA, supabase

    results = {
        "supabase_url": SUPABASE_URL,
        "supabase_key_prefix": SUPABASE_KEY[:20] + "..." if SUPABASE_KEY else None,
        "supabase_key_length": len(SUPABASE_KEY) if SUPABASE_KEY else 0,
        "schema": SCHEMA,
        "tests": {}
    }

    # Test 1: consulta al esquema público (sin schema custom)
    try:
        resp = supabase.from_('lista_personal').select('*').limit(1).execute()
        results["tests"]["public_schema_query"] = {
            "status": "ok",
            "rows": len(resp.data or [])
        }
    except Exception as e:
        results["tests"]["public_schema_query"] = {
            "status": "error",
            "error": str(e),
            "traceback": traceback.format_exc()
        }

    # Test 2: consulta al esquema jcf
    try:
        resp = supabase.schema(SCHEMA).from_('lista_personal').select('*').limit(1).execute()
        results["tests"]["custom_schema_query"] = {
            "status": "ok",
            "rows": len(resp.data or [])
        }
    except Exception as e:
        results["tests"]["custom_schema_query"] = {
            "status": "error",
            "error": str(e),
            "traceback": traceback.format_exc()
        }

    return jsonify(results)

@app.route('/api/encargos', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_encargos():
    if request.method == 'GET':
        result = sb_select('ENCARGOS')
        # Oracle: result = execute_query("SELECT * FROM ENCARGOS")
        return jsonify(result)

    elif request.method == 'POST':
        data = request.json
        sb_insert('ENCARGOS', data)
        # Oracle: execute_query("INSERT INTO ENCARGOS (...) VALUES (...)", data, is_select=False)
        return jsonify({"status": "success"}), 201

    elif request.method == 'PUT':
        data = request.json
        sb_update('ENCARGOS', data, {'CODIGOPR': data.get('CODIGOPR')})
        # Oracle: execute_query("UPDATE ENCARGOS SET ... WHERE CODIGOPR=:CODIGOPR", data, is_select=False)
        return jsonify({"status": "success"})

    elif request.method == 'DELETE':
        codigopr = request.args.get('codigopr')
        sb_delete('ENCARGOS', {'CODIGOPR': codigopr})
        # Oracle: execute_query("DELETE FROM ENCARGOS WHERE CODIGOPR=:codigopr", {"codigopr": codigopr}, is_select=False)
        return jsonify({"status": "success"})

@app.route('/api/personal/bulk-location', methods=['POST'])
def bulk_update_personal_location():
    data = request.json
    ref_pers = data.get('ref_pers', [])
    new_location = data.get('new_location')

    if not ref_pers or not new_location:
        return jsonify({"error": "Se requieren IDs de personal y la nueva ubicación"}), 400

    try:
        for ref_per in ref_pers:
            sb_update('LISTA_PERSONAL', {'REF_UBI': new_location}, {'REF_PER': ref_per})
            # Oracle: execute_query("UPDATE LISTA_PERSONAL SET REF_UBI=:new_location WHERE REF_PER=:ref_per", ..., is_select=False)
        return jsonify({"status": "success", "message": f"Ubicación actualizada para {len(ref_pers)} personas"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/personal', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_personal():
    if request.method == 'GET':
        result = sb_select('LISTA_PERSONAL')
        # Oracle: result = execute_query("SELECT * FROM LISTA_PERSONAL")
        return jsonify(result)

    elif request.method == 'POST':
        data = request.json
        sb_insert('LISTA_PERSONAL', data)
        # Oracle: execute_query("INSERT INTO LISTA_PERSONAL (...) VALUES (...)", data, is_select=False)
        return jsonify({"status": "success"}), 201

    elif request.method == 'PUT':
        data = request.json
        sb_update('LISTA_PERSONAL', data, {'REF_PER': data.get('REF_PER')})
        # Oracle: execute_query("UPDATE LISTA_PERSONAL SET ... WHERE REF_PER=:REF_PER", data, is_select=False)
        return jsonify({"status": "success"})

    elif request.method == 'DELETE':
        ref_per = request.args.get('ref_per')
        sb_delete('LISTA_PERSONAL', {'REF_PER': ref_per})
        # Oracle: execute_query("DELETE FROM LISTA_PERSONAL WHERE REF_PER=:ref_per", {"ref_per": ref_per}, is_select=False)
        return jsonify({"status": "success"})

@app.route('/api/personal-proyectos', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_personal_proyectos():
    codigopr = request.args.get('codigopr')
    
    if request.method == 'GET':
        if codigopr:
            assignments = sb_select('PERSONAL_PROYECTOS', {'CODIGOPR': codigopr})
        else:
            assignments = sb_select('PERSONAL_PROYECTOS')
        # Oracle path removed:
        # query = "SELECT pp.*, p.NOMBRE... FROM PERSONAL_PROYECTOS pp JOIN LISTA_PERSONAL p ... WHERE pp.CODIGOPR = :codigopr"
        try:
            personal_rows = sb_select('LISTA_PERSONAL')
            personal_map = {str(p.get('REF_PER')): p for p in personal_rows}
            result = []
            for a in assignments:
                row = {**a}
                person = personal_map.get(str(a.get('REF_PER')))
                if person:
                    row['NOMBRE'] = person.get('NOMBRE')
                    row['APELLIDO1'] = person.get('APELLIDO1')
                    row['APELLIDO2'] = person.get('APELLIDO2')
                    row['PERFIL'] = person.get('PERFIL')
                result.append(row)
        except Exception as e:
            logging.error(f"[PERSONAL_PROYECTOS] Error joining LISTA_PERSONAL: {str(e)}")
            result = assignments
        return jsonify(result)

    elif request.method == 'POST':
        data = request.json
        sb_insert('PERSONAL_PROYECTOS', data)
        # Oracle: execute_query("INSERT INTO PERSONAL_PROYECTOS (...) VALUES (...)", data, is_select=False)
        return jsonify({"status": "success"}), 201

    elif request.method == 'PUT':
        data = request.json
        sb_update('PERSONAL_PROYECTOS', data, {'REF_PER': data.get('REF_PER'), 'CODIGOPR': data.get('CODIGOPR')})
        # Oracle: execute_query("UPDATE PERSONAL_PROYECTOS SET ... WHERE REF_PER=:REF_PER AND CODIGOPR=:CODIGOPR", data, is_select=False)
        return jsonify({"status": "success"})

    elif request.method == 'DELETE':
        ref_per = request.args.get('ref_per')
        codigopr = request.args.get('codigopr')
        sb_delete('PERSONAL_PROYECTOS', {'REF_PER': ref_per, 'CODIGOPR': codigopr})
        # Oracle: execute_query("DELETE FROM PERSONAL_PROYECTOS WHERE REF_PER=:ref_per AND CODIGOPR=:codigopr", ..., is_select=False)
        return jsonify({"status": "success"})

@app.route('/api/ubicacion', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_ubicacion():
    if request.method == 'GET':
        result = sb_select('UBICACION', order='REF_UBI')
        # Oracle: result = execute_query("SELECT * FROM UBICACION ORDER BY REF_UBI")
        return jsonify(result)

    elif request.method == 'POST':
        data = request.json
        sb_insert('UBICACION', data)
        # Oracle: execute_query("INSERT INTO UBICACION (REF_UBI, A_LUGAR) VALUES (:REF_UBI, :A_LUGAR)", data, is_select=False)
        return jsonify({"status": "success"}), 201

    elif request.method == 'PUT':
        data = request.json
        sb_update('UBICACION', data, {'REF_UBI': data.get('REF_UBI')})
        # Oracle: execute_query("UPDATE UBICACION SET A_LUGAR=:A_LUGAR WHERE REF_UBI=:REF_UBI", data, is_select=False)
        return jsonify({"status": "success"})

    elif request.method == 'DELETE':
        ref_ubi = request.args.get('ref_ubi')
        sb_delete('UBICACION', {'REF_UBI': ref_ubi})
        # Oracle: execute_query("DELETE FROM UBICACION WHERE REF_UBI=:ref_ubi", {"ref_ubi": ref_ubi}, is_select=False)
        return jsonify({"status": "success"})

@app.route('/api/schema', methods=['GET'])
def get_schema():
    # Oracle path removed - always PostgreSQL/Supabase
    # Oracle query was: SELECT table_name, column_name FROM user_tab_columns WHERE table_name IN (...) ...
    schema_name = config.PG_SCHEMA or 'jcf'
    from supabase_client import supabase, SCHEMA
    try:
        resp = supabase.schema(SCHEMA).rpc('get_schema_columns', {}).execute()
    except Exception:
        pass
    # Fallback: build schema from information_schema via psycopg2
    from db import execute_query as pg_query
    query = f"""
        SELECT table_name, column_name
        FROM information_schema.columns
        WHERE table_schema = '{schema_name}'
          AND LOWER(table_name) IN ('lista_personal', 'encargos', 'personal_proyectos', 'ubicacion')
        ORDER BY table_name, ordinal_position
    """
    result = pg_query(query)
    schema = {}
    for row in result:
        table = (row.get('TABLE_NAME') or row.get('table_name')).upper()
        column = (row.get('COLUMN_NAME') or row.get('column_name')).upper()
        if table not in schema:
            schema[table] = []
        schema[table].append(column)
    return jsonify(schema)

def parse_iso_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val
    try:
        val_str = str(val).split('+')[0].rstrip('Z')
        return datetime.datetime.fromisoformat(val_str)
    except Exception:
        try:
            return datetime.datetime.strptime(val_str.split('T')[0], "%Y-%m-%d")
        except Exception:
            return val

@app.route('/api/backup/export', methods=['GET'])
def backup_export():
    try:
        tables = ['UBICACION', 'ENCARGOS', 'LISTA_PERSONAL', 'PERSONAL_PROYECTOS', 'VACACIONES', 'FESTIVOS']
        backup_data = {}
        for table in tables:
            try:
                if is_postgres():
                    backup_data[table] = sb_select(table)
                else:
                    query = f"SELECT * FROM {table}"
                    backup_data[table] = execute_query(query)
            except Exception as e:
                logging.warning(f"No se pudo exportar la tabla {table} (puede que no exista): {str(e)}")
                backup_data[table] = []
        return jsonify({
            "status": "success",
            "timestamp": datetime.datetime.now().isoformat(),
            "data": backup_data
        })
    except Exception as e:
        logging.error(f"Error exporting backup: {str(e)}")
        return jsonify({"error": f"Error al exportar copia de seguridad: {str(e)}"}), 500

@app.route('/api/backup/import', methods=['POST'])
def backup_import():
    payload = request.json
    if not payload or 'data' not in payload:
        return jsonify({"error": "Datos de copia de seguridad no proporcionados"}), 400
    
    backup_data = payload['data']
    # Delete order (dependents first)
    tables_to_delete = ['VACACIONES', 'PERSONAL_PROYECTOS', 'LISTA_PERSONAL', 'ENCARGOS', 'FESTIVOS', 'UBICACION']
    
    try:
        # Step 1: Clear current tables
        for table in tables_to_delete:
            try:
                if is_postgres():
                    sb_delete_all(table)
                else:
                    execute_query(f"DELETE FROM {table}", is_select=False)
            except Exception as e:
                logging.warning(f"No se pudo vaciar la tabla {table} (puede que no exista): {str(e)}")
            
        inserted_counts = {}
        
        # 1. UBICACION
        if 'UBICACION' in backup_data and backup_data['UBICACION']:
            try:
                if is_postgres():
                    sb_insert('UBICACION', backup_data['UBICACION'])
                else:
                    for row in backup_data['UBICACION']:
                        execute_query("INSERT INTO UBICACION (REF_UBI, A_LUGAR) VALUES (:REF_UBI, :A_LUGAR)", row, is_select=False)
                inserted_counts['UBICACION'] = len(backup_data['UBICACION'])
            except Exception as e:
                logging.warning(f"No se pudo importar UBICACION: {str(e)}")
            
        # 3. ENCARGOS
        if 'ENCARGOS' in backup_data and backup_data['ENCARGOS']:
            try:
                rows = []
                for row in backup_data['ENCARGOS']:
                    clean_row = {**row}
                    for f in ['INICIO', 'FIN', 'FIN_REAL']:
                        clean_row[f] = parse_iso_date(clean_row.get(f))
                        if isinstance(clean_row[f], (datetime.datetime, datetime.date)):
                            clean_row[f] = clean_row[f].strftime('%Y-%m-%d')
                    rows.append(clean_row)
                sb_insert('ENCARGOS', rows)
                # Oracle: execute_query("INSERT INTO ENCARGOS ...", clean_row, is_select=False)
                inserted_counts['ENCARGOS'] = len(backup_data['ENCARGOS'])
            except Exception as e:
                logging.warning(f"No se pudo importar ENCARGOS: {str(e)}")
            
        # 4. LISTA_PERSONAL
        if 'LISTA_PERSONAL' in backup_data and backup_data['LISTA_PERSONAL']:
            try:
                rows = []
                for row in backup_data['LISTA_PERSONAL']:
                    clean_row = {**row}
                    for f in ['BAJA', 'INCORPORACION', 'F_CONTRATO']:
                        clean_row[f] = parse_iso_date(clean_row.get(f))
                        if isinstance(clean_row[f], (datetime.datetime, datetime.date)):
                            clean_row[f] = clean_row[f].strftime('%Y-%m-%d')
                    rows.append(clean_row)
                sb_insert('LISTA_PERSONAL', rows)
                # Oracle: execute_query("INSERT INTO LISTA_PERSONAL ...", clean_row, is_select=False)
                inserted_counts['LISTA_PERSONAL'] = len(backup_data['LISTA_PERSONAL'])
            except Exception as e:
                logging.warning(f"No se pudo importar LISTA_PERSONAL: {str(e)}")
            
        # 5. PERSONAL_PROYECTOS
        if 'PERSONAL_PROYECTOS' in backup_data and backup_data['PERSONAL_PROYECTOS']:
            try:
                rows = []
                for row in backup_data['PERSONAL_PROYECTOS']:
                    clean_row = {**row}
                    for f in ['ALTA', 'BAJA']:
                        clean_row[f] = parse_iso_date(clean_row.get(f))
                        if isinstance(clean_row[f], (datetime.datetime, datetime.date)):
                            clean_row[f] = clean_row[f].strftime('%Y-%m-%d')
                    rows.append(clean_row)
                sb_insert('PERSONAL_PROYECTOS', rows)
                # Oracle: execute_query("INSERT INTO PERSONAL_PROYECTOS ...", clean_row, is_select=False)
                inserted_counts['PERSONAL_PROYECTOS'] = len(backup_data['PERSONAL_PROYECTOS'])
            except Exception as e:
                logging.warning(f"No se pudo importar PERSONAL_PROYECTOS: {str(e)}")

        # 6. VACACIONES
        if 'VACACIONES' in backup_data and backup_data['VACACIONES']:
            try:
                rows = []
                for row in backup_data['VACACIONES']:
                    clean_row = {**row}
                    clean_row['FECHA_DESDE'] = parse_iso_date(clean_row.get('FECHA_DESDE'))
                    clean_row['FECHA_HASTA'] = parse_iso_date(clean_row.get('FECHA_HASTA'))
                    if isinstance(clean_row['FECHA_DESDE'], (datetime.datetime, datetime.date)):
                        clean_row['FECHA_DESDE'] = clean_row['FECHA_DESDE'].strftime('%Y-%m-%d')
                    if isinstance(clean_row['FECHA_HASTA'], (datetime.datetime, datetime.date)):
                        clean_row['FECHA_HASTA'] = clean_row['FECHA_HASTA'].strftime('%Y-%m-%d')
                    rows.append(clean_row)
                sb_insert('VACACIONES', rows)
                # Oracle: execute_query("INSERT INTO VACACIONES ...", row, is_select=False)
                inserted_counts['VACACIONES'] = len(backup_data['VACACIONES'])
            except Exception as e:
                logging.warning(f"No se pudo importar VACACIONES: {str(e)}")
            
        # 7. FESTIVOS
        if 'FESTIVOS' in backup_data and backup_data['FESTIVOS']:
            try:
                rows = []
                for row in backup_data['FESTIVOS']:
                    clean_row = {**row}
                    fecha = parse_iso_date(clean_row.get('FECHA') or clean_row.get('fecha'))
                    if isinstance(fecha, (datetime.datetime, datetime.date)):
                        fecha_str = fecha.strftime('%Y-%m-%d')
                    else:
                        fecha_str = str(fecha) if fecha is not None else None
                    rows.append({
                        'YEAR': int(clean_row.get('YEAR') or clean_row.get('year')) if (clean_row.get('YEAR') or clean_row.get('year')) is not None else None,
                        'REF_UBI': int(clean_row.get('REF_UBI') or clean_row.get('ref_ubi')) if (clean_row.get('REF_UBI') or clean_row.get('ref_ubi')) is not None else None,
                        'FECHA': fecha_str,
                        'DESCRIPCION': clean_row.get('DESCRIPCION') or clean_row.get('descripcion')
                    })
                sb_insert('FESTIVOS', rows)
                # Oracle: execute_query("INSERT INTO FESTIVOS ...", row, is_select=False)
                inserted_counts['FESTIVOS'] = len(backup_data['FESTIVOS'])
            except Exception as e:
                logging.warning(f"No se pudo importar FESTIVOS: {str(e)}")
            
        return jsonify({
            "status": "success",
            "message": "Datos importados con éxito",
            "inserted_counts": inserted_counts
        })
    except Exception as e:
        logging.error(f"Error importing backup: {str(e)}")
        return jsonify({"error": f"Error al importar copia de seguridad: {str(e)}"}), 500

@app.route('/api/query', methods=['POST'])
def execute_dynamic_query():
    data = request.json
    sql = data.get('sql')
    if not sql:
        return jsonify({"error": "SQL query is required"}), 400
    
    # Basic security check - only allow SELECT
    if not sql.strip().upper().startswith('SELECT'):
        return jsonify({"error": "Only SELECT queries are allowed"}), 403
        
    try:
        result = execute_query(sql)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

def check_and_create_vacaciones_table():
    # Only postgres now, no need to create dynamically
    return
    # ============================================================
    # ORACLE TABLE CREATION - COMMENTED OUT (not in use)
    # ============================================================
    # try:
    #     check_query = "SELECT 1 FROM VACACIONES WHERE ROWNUM = 1"
    #     execute_query(check_query)
    # except Exception:
    #     logging.info("Creating VACACIONES table and sequence for Oracle...")
    #     ...


def check_and_create_festivos_table():
    from flask import g
    db_type = getattr(g, 'db_type', 'oracle')
    if db_type == 'postgres':
        # Supabase debe tener la tabla FESTIVOS creada previamente.
        return
    else:
        try:
            execute_query("SELECT 1 FROM FESTIVOS WHERE ROWNUM = 1")
        except Exception:
            logging.info("Creating FESTIVOS table and sequence for Oracle...")
            try:
                execute_query("CREATE SEQUENCE SEQ_FESTIVOS START WITH 1 INCREMENT BY 1", is_select=False)
            except Exception as seq_err:
                logging.warning(f"Sequence SEQ_FESTIVOS creation skipped: {seq_err}")
            create_sql = """
                CREATE TABLE FESTIVOS (
                    ID_FESTIVO NUMBER PRIMARY KEY,
                    YEAR NUMBER(4) NOT NULL,
                    REF_UBI NUMBER,
                    FECHA DATE NOT NULL,
                    DESCRIPCION VARCHAR2(255)
                )
            """
            execute_query(create_sql, is_select=False)
            
@app.route('/api/vacaciones', methods=['GET'])
def get_vacaciones():
    check_and_create_vacaciones_table()
    
    def clean_arg(val):
        if val in ('null', 'undefined', '', 'NaN'):
            return None
        return val

    ref_per = clean_arg(request.args.get('ref_per'))
    year = clean_arg(request.args.get('year'))
    origen_fichero = clean_arg(request.args.get('origen_fichero'))
    codigopr = clean_arg(request.args.get('codigopr'))
    
    logging.debug(f"[VACACIONES] Request args: ref_per={ref_per}, year={year}, origen_fichero={origen_fichero}, codigopr={codigopr}")

    try:
        vacations = sb_select('VACACIONES')
        logging.debug(f"[VACACIONES] rows retrieved: {len(vacations)}")
        if ref_per:
            vacations = [row for row in vacations if str(row.get('REF_PER')) == str(ref_per)]
        if year:
            vacations = [row for row in vacations if row.get('FECHA_DESDE') and str(row.get('FECHA_DESDE')).startswith(str(year))]
        if origen_fichero:
            vacations = [row for row in vacations if str(row.get('ORIGEN_FICHERO')) == str(origen_fichero)]
        if codigopr:
            proyectos = sb_select('PERSONAL_PROYECTOS', {'CODIGOPR': codigopr})
            ref_pers = {str(row.get('REF_PER')) for row in proyectos}
            vacations = [row for row in vacations if str(row.get('REF_PER')) in ref_pers]

        personal_rows = sb_select('LISTA_PERSONAL')
        personal_map = {str(person.get('REF_PER')): person for person in personal_rows}

        result = []
        for vac in vacations:
            row = {**vac}
            person = personal_map.get(str(vac.get('REF_PER')))
            if person:
                row['NOMBRE'] = person.get('NOMBRE')
                row['APELLIDO1'] = person.get('APELLIDO1')
                row['APELLIDO2'] = person.get('APELLIDO2')
                row['PERFIL'] = person.get('PERFIL')
                row['USUARIO'] = person.get('USUARIO')
            result.append(row)
        return jsonify(result)
    except Exception as e:
        import traceback
        logging.error(f"[VACACIONES] ERROR: {str(e)}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/vacaciones/ficheros', methods=['GET'])
def get_vacaciones_ficheros():
    check_and_create_vacaciones_table()
    try:
        return jsonify(sb_select_distinct('VACACIONES', 'ORIGEN_FICHERO'))
        # Oracle: result = execute_query("SELECT DISTINCT ORIGEN_FICHERO FROM VACACIONES ...")
        # return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vacaciones/import', methods=['POST'])
def import_vacaciones():
    check_and_create_vacaciones_table()
    data = request.json
    if not isinstance(data, list):
        return jsonify({"error": "Se requiere una lista de registros de vacaciones"}), 400
        
    inserted_count = 0
    try:
        rows = []
        for row in data:
            fecha_desde = parse_iso_date(row.get('fecha_desde'))
            fecha_hasta = parse_iso_date(row.get('fecha_hasta'))
            if not fecha_desde or not fecha_hasta:
                continue
            rows.append({
                'REF_PER': int(row['ref_per']),
                'DURACION': float(row.get('duracion')) if row.get('duracion') is not None else None,
                'FECHA_DESDE': fecha_desde.strftime('%Y-%m-%d') if isinstance(fecha_desde, (datetime.datetime, datetime.date)) else str(fecha_desde).split('T')[0],
                'FECHA_HASTA': fecha_hasta.strftime('%Y-%m-%d') if isinstance(fecha_hasta, (datetime.datetime, datetime.date)) else str(fecha_hasta).split('T')[0],
                'PARTICION_NUM': int(row['particion_num']) if row.get('particion_num') is not None else None,
                'ORIGEN_FICHERO': row.get('origen_fichero', 'Carga manual')
            })
        sb_insert('VACACIONES', rows)
        inserted_count = len(rows)
        return jsonify({"status": "success", "message": f"{inserted_count} periodos de vacaciones importados correctamente"}), 201
        
        # Oracle path removed
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vacaciones', methods=['PUT'])
def update_vacacion():
    check_and_create_vacaciones_table()
    data = request.json or {}
    id_vacacion = data.get('id')
    if not id_vacacion:
        return jsonify({"error": "Se requiere el ID de vacaciones"}), 400

    duracion = data.get('duracion')
    fecha_desde = parse_iso_date(data.get('fecha_desde'))
    fecha_hasta = parse_iso_date(data.get('fecha_hasta'))
    particion_num = data.get('particion_num')
    origen_fichero = data.get('origen_fichero')

    if not fecha_desde or not fecha_hasta:
        return jsonify({"error": "Se requiere fecha desde y fecha hasta en formato YYYY-MM-DD"}), 400

    if isinstance(fecha_desde, (datetime.datetime, datetime.date)):
        fecha_desde = fecha_desde.strftime('%Y-%m-%d')
    if isinstance(fecha_hasta, (datetime.datetime, datetime.date)):
        fecha_hasta = fecha_hasta.strftime('%Y-%m-%d')

    try:
        sb_update('VACACIONES', {
            'DURACION': float(duracion) if duracion is not None and duracion != '' else None,
            'FECHA_DESDE': fecha_desde,
            'FECHA_HASTA': fecha_hasta,
            'PARTICION_NUM': int(particion_num) if particion_num is not None and particion_num != '' else None,
            'ORIGEN_FICHERO': origen_fichero
        }, {'ID_VACACION': int(id_vacacion)})
        return jsonify({"status": "success", "message": "Periodo de vacaciones actualizado correctamente"})
        # Oracle path removed
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vacaciones', methods=['DELETE'])
def delete_vacaciones():
    check_and_create_vacaciones_table()
    id_vacacion = request.args.get('id')
    origen_fichero = request.args.get('origen_fichero')

    if not id_vacacion and not origen_fichero:
        return jsonify({"error": "Se requiere el ID de vacaciones o el nombre del fichero de origen"}), 400

    try:
        if id_vacacion:
            sb_delete('VACACIONES', {'ID_VACACION': int(id_vacacion)})
            message = "Periodo de vacaciones eliminado correctamente"
        else:
            sb_delete('VACACIONES', {'ORIGEN_FICHERO': origen_fichero})
            message = f"Todas las vacaciones del fichero '{origen_fichero}' han sido eliminadas"
        return jsonify({"status": "success", "message": message})
        # Oracle path removed
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/festivos', methods=['GET', 'POST', 'PUT', 'DELETE'])
def manage_festivos():
    from flask import g
    check_and_create_festivos_table()

    if request.method == 'GET':
        year = request.args.get('year')
        ref_ubi = request.args.get('ref_ubi')
        filters = {}
        if year:
            filters['YEAR'] = int(year)
        if ref_ubi:
            filters['REF_UBI'] = int(ref_ubi)
        try:
            return jsonify(sb_select('FESTIVOS', filters))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    if request.method == 'POST':
        data = request.json or {}
        year = int(data.get('year'))
        ref_ubi = data.get('ref_ubi')
        fecha = data.get('fecha')
        descripcion = data.get('descripcion')
        if not year or not fecha:
            return jsonify({"error": "Se requieren 'year' y 'fecha'"}), 400
        try:
            sb_insert('FESTIVOS', [{
                'YEAR': year,
                'REF_UBI': int(ref_ubi) if ref_ubi is not None else None,
                'FECHA': fecha,
                'DESCRIPCION': descripcion
            }])
            return jsonify({"status": "success"}), 201
            # Oracle path removed
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    if request.method == 'PUT':
        data = request.json or {}
        id_festivo = data.get('id_festivo') or data.get('id')
        if not id_festivo:
            return jsonify({"error": "Se requiere 'id_festivo'"}), 400
        year = data.get('year')
        ref_ubi = data.get('ref_ubi')
        fecha = data.get('fecha')
        descripcion = data.get('descripcion')
        try:
            sb_update('FESTIVOS', {
                'YEAR': int(year) if year is not None else None,
                'REF_UBI': int(ref_ubi) if ref_ubi is not None else None,
                'FECHA': fecha,
                'DESCRIPCION': descripcion
            }, {'ID_FESTIVO': int(id_festivo)})
            return jsonify({"status": "success"})
            # Oracle path removed
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    if request.method == 'DELETE':
        id_festivo = request.args.get('id') or request.args.get('id_festivo')
        if not id_festivo:
            return jsonify({"error": "Se requiere 'id' param para eliminar"}), 400
        try:
            sb_delete('FESTIVOS', {'ID_FESTIVO': int(id_festivo)})
            return jsonify({"status": "success"})
            # Oracle path removed
        except Exception as e:
            return jsonify({"error": str(e)}), 500


@app.route('/api/export-calendar', methods=['POST'])
def export_calendar():
    """Export calendar with styling to XLSX using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        
        data = request.json
        months = data.get('months', [])  # List of {monthYear, days}
        employees = data.get('employees', [])  # List of employee names
        cellColors = data.get('cellColors', {})  # {cellRef: colorHex}
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Calendario'
        
        # Row 1: Month headers
        ws.append(['Empleado'] + [m['monthYear'].split()[0] for m in months])
        
        # Row 2: Day numbers
        dayRow = ['']
        for month in months:
            for day in month.get('days', []):
                dayRow.append(day.get('number', ''))
        ws.append(dayRow)
        
        # Employee data rows (starts at row 3)
        for emp in employees:
            row = [emp]
            # Add empty cells for each day
            totalDays = sum(len(m.get('days', [])) for m in months)
            row.extend([''] * totalDays)
            ws.append(row)
        
        # Apply colors
        colorMap = {
            'EF4444': PatternFill(start_color='FFEF4444', end_color='FFEF4444', fill_type='solid'),
            '60A5FA': PatternFill(start_color='FF60A5FA', end_color='FF60A5FA', fill_type='solid'),
            'C084FC': PatternFill(start_color='FFC084FC', end_color='FFC084FC', fill_type='solid'),
            'A7F3D0': PatternFill(start_color='FFA7F3D0', end_color='FFA7F3D0', fill_type='solid')
        }
        
        for cellRef, colorHex in cellColors.items():
            try:
                cell = ws[cellRef]
                if colorHex in colorMap:
                    cell.fill = colorMap[colorHex]
            except:
                pass  # Skip invalid cell refs
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        for i in range(2, 100):
            ws.column_dimensions[chr(64 + i)].width = 4
        
        # Freeze panes
        ws.freeze_panes = 'B3'
        
        # Export to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'calendario_{data.get("year", 2026)}.xlsx'
        )
    except Exception as e:
        logging.error(f"Export error: {str(e)}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, port=config.PORT, host='0.0.0.0')

